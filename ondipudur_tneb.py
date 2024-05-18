

import pandas as pd
from openpyxl import load_workbook
import seaborn as sns
import matplotlib.pyplot as plt

df = pd.read_excel('/kaggle/input/eb-tariff/OPR CC Details New.xlsx')

wb = load_workbook('/kaggle/input/eb-tariff/OPR CC Details New.xlsx')
sheet = wb.active

merged_ranges_copy = sheet.merged_cells.ranges.copy()

for merged_range in merged_ranges_copy:
    merged_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
    if merged_value is not None:
        # Convert values to integers before unmerging
        start_row, start_col, end_row, end_col = (
            merged_range.min_row, merged_range.min_col,
            merged_range.max_row, merged_range.max_col
        )
        sheet.unmerge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

        for row in sheet.iter_rows(min_row=start_row, max_row=end_row,
                                   min_col=start_col, max_col=end_col):
            for cell in row:
                cell.value = merged_value

output_excel_file_path = '/kaggle/working/opr_final.xlsx'
wb.save(output_excel_file_path)

# Read the modified Excel file into a Pandas DataFrame
df = pd.read_excel('/kaggle/working/opr_final.xlsx', header=1)

df.head(4)

df.isna().sum()

df = df.dropna(subset=['SC number'])

df['CC ISSUSED BY (DTCP/Corpn etc'] = df['CC ISSUSED BY (DTCP/Corpn etc'].fillna(method='ffill')

grouped_df = df.groupby('CC ISSUSED BY (DTCP/Corpn etc')

df['CC ISSUSED BY (DTCP/Corpn etc'].nunique()

df['Tariff'].unique()

fig, axes = plt.subplots(nrows=7, ncols=6, figsize=(25, 25))

# Flatten the 6x6 grid into a 1D array for easy iteration
axes = axes.flatten()

# Plot pie charts for each group based on the 'Category' column
for i, (name, group) in enumerate(grouped_df):
    value_counts = group['Tariff'].value_counts()

    # Use the current subplot from the flattened array
    ax = axes[i]

    ax.pie(value_counts, labels=value_counts.index, autopct='%1.1f%%', startangle=90)
    ax.set_title(name)
    ax.legend(value_counts.index, title='Tariff', loc='upper right', bbox_to_anchor=(1.2, 1))

# Hide any remaining empty subplots in the grid
for ax in axes[len(grouped_df):]:
    ax.axis('off')

# Adjust layout to prevent overlapping
plt.tight_layout()

# Save the figure to a file
plt.savefig('ondipudur_pie.png', bbox_inches='tight')

column_name = 'Tariff'
row = 'CC ISSUSED BY (DTCP/Corpn etc'

# Create countplot
plt.figure(figsize=(20, 12))
sns.countplot(x=row, hue=column_name, data=df)
plt.xticks(rotation=45, ha='right')

# for label in ax.xaxis.get_ticklabels()[::10]:
#     label.set_visible(False)


# Set labels and title
plt.xlabel(row)
plt.ylabel('Count')
plt.yscale('log')
plt.title(f'Tariff issued by {row}')
plt.tight_layout()
plt.savefig('ondipudur_tariff_countplot_log.png')

plt.show()

df.info()

df.to_csv('ondipudur.csv', index=False)

